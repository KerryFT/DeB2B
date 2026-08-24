from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.infrastructure.models import EvidenceSpan


@dataclass(frozen=True, slots=True)
class ExtractedField:
    name: str
    value: Any
    sheet: str
    cell_range: str
    quote: str


CRITICAL_FIELDS = ("invoice_number", "customer_code", "amount", "due_date")


def extract_invoice_xlsx(content: bytes, *, row_number: int = 2) -> list[ExtractedField]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    headers = {str(cell.value or "").strip(): cell.column for cell in next(sheet.iter_rows())}
    missing = set(CRITICAL_FIELDS) - headers.keys()
    if missing:
        raise ValueError(f"missing critical columns: {', '.join(sorted(missing))}")
    fields = []
    for name in CRITICAL_FIELDS:
        cell = sheet.cell(row=row_number, column=headers[name])
        fields.append(
            ExtractedField(
                name=name,
                value=cell.value,
                sheet=sheet.title,
                cell_range=cell.coordinate,
                quote=str(cell.value),
            )
        )
    return fields


def persist_evidence(
    session: Session,
    *,
    tenant_id: UUID,
    document_id: UUID,
    fields: list[ExtractedField],
) -> list[EvidenceSpan]:
    spans = [
        EvidenceSpan(
            tenant_id=tenant_id,
            document_id=document_id,
            field_name=field.name,
            page=None,
            sheet=field.sheet,
            cell_range=field.cell_range,
            polygon=None,
            quote=field.quote,
        )
        for field in fields
    ]
    session.add_all(spans)
    session.flush()
    return spans
