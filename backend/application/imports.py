from __future__ import annotations

import csv
import hashlib
import io
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field, ValidationError, field_validator

REQUIRED_COLUMNS = {
    "invoice_number",
    "customer_code",
    "customer_name",
    "issue_date",
    "due_date",
    "amount",
    "currency",
    "outstanding_amount",
}


class InvoiceImportRow(BaseModel):
    invoice_number: str = Field(min_length=1, max_length=100)
    customer_code: str = Field(min_length=1, max_length=100)
    customer_name: str = Field(min_length=1, max_length=300)
    tax_id: str | None = None
    issue_date: date
    due_date: date
    amount: Decimal = Field(ge=0)
    currency: str = Field(pattern=r"^[A-Za-z]{3}$")
    outstanding_amount: Decimal = Field(ge=0)
    account_owner: str | None = None
    status: str | None = None

    @field_validator("currency")
    @classmethod
    def normalize_currency(cls, value: str) -> str:
        return value.upper()

    @field_validator("outstanding_amount")
    @classmethod
    def finite_amount(cls, value: Decimal) -> Decimal:
        if not value.is_finite():
            raise ValueError("amount must be finite")
        return value

    def fingerprint(self) -> str:
        source = "|".join(
            [self.customer_code, self.invoice_number, self.issue_date.isoformat(), self.currency]
        )
        return hashlib.sha256(source.encode()).hexdigest()


class RowError(BaseModel):
    row_number: int
    errors: list[dict[str, Any]]


class ImportPreview(BaseModel):
    valid: list[InvoiceImportRow]
    invalid: list[RowError]
    columns: list[str]


def _records_from_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    columns = list(reader.fieldnames or [])
    return columns, list(reader)


def _records_from_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    columns = [str(value or "").strip() for value in next(iterator)]
    records = [dict(zip(columns, row, strict=True)) for row in iterator]
    return columns, records


def preview_import(content: bytes, filename: str) -> ImportPreview:
    columns, records = (
        _records_from_xlsx(content)
        if filename.lower().endswith(".xlsx")
        else _records_from_csv(content)
    )
    missing = sorted(REQUIRED_COLUMNS - set(columns))
    if missing:
        raise ValueError(f"missing required columns: {', '.join(missing)}")
    valid: list[InvoiceImportRow] = []
    invalid: list[RowError] = []
    for offset, record in enumerate(records, start=2):
        try:
            row = InvoiceImportRow.model_validate(record)
            if row.outstanding_amount > row.amount:
                raise ValueError("outstanding_amount exceeds amount")
            valid.append(row)
        except (ValidationError, ValueError, InvalidOperation) as exc:
            errors = exc.errors() if isinstance(exc, ValidationError) else [{"msg": str(exc)}]
            invalid.append(RowError(row_number=offset, errors=errors))
    return ImportPreview(valid=valid, invalid=invalid, columns=columns)
